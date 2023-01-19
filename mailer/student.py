import openpyxl as opx
import csv


class Student:
    def __init__(self, name:str, email: str, group: str):
        self.email = email
        self.group = group
        self.name = name

    @staticmethod
    def read_students(path: str):
        wb = opx.load_workbook(filename=path)
        ws = wb.active
        students = []
        for i in range(1, 326 + 1):
            email = ws[f'C{i}'].value.strip()
            group = ws[f'K{i}'].value.strip()[3:]

            students.append(Student(email, group))
        return students

    @staticmethod
    def read_students_csv(path: str = 'students.csv'):
        students = []
        with open(path) as csvfile:
            reader = csv.reader(csvfile)
            for row in reader:
                email = row[2]
                name = row[1]
                if email == "Адрес электронной почты":
                    continue
                group = row[10][3:]
                students.append(Student(name, email, group))
        return students
